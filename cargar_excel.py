import os
import pandas as pd
from decimal import Decimal
from datetime import datetime
from utils import seguro_decimal
from flask import Blueprint, render_template, request, flash, redirect, url_for, send_file
from werkzeug.utils import secure_filename
from flask_login import login_required, current_user
from functools import wraps
from flask import abort
from models import (db, Cliente, Producto, Proveedor, Venta, DetalleVenta,
                    Compra, CompraDetalle, CuentaPorPagar, HistorialPago, TasaBCV,
                    MovimientoCaja, AuditoriaInventario)
from routes.contabilidad import registrar_asiento

cargar_bp = Blueprint('cargar', __name__)

UPLOAD_FOLDER = os.path.join(os.getcwd(), 'importar_aqui')
PLANTILLAS_FOLDER = os.path.join(os.getcwd(), 'static', 'plantillas')

# ============================================================
# 🔒 DECORADOR DE SEGURIDAD
# ============================================================
def solo_admin(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            flash("⚠️ Debes iniciar sesión primero.", "warning")
            return redirect(url_for('auth.ingresar'))
        if current_user.role not in ['admin', 'supervisor']:
            flash("🚫 No tienes permiso para acceder a esta sección.", "danger")
            abort(403)
        return f(*args, **kwargs)
    return decorated_function

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xls', 'xlsx'}

def guardar_archivo(file):
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    filename = secure_filename(file.filename)
    ruta = os.path.join(UPLOAD_FOLDER, filename)
    file.save(ruta)
    return ruta

# ============================================================
#   PANEL PRINCIPAL DE CARGA 🔒
# ============================================================
@cargar_bp.route('/cargar')
@login_required
@solo_admin
def panel_cargar():
    return render_template('cargar_excel.html')

# ============================================================
#   1. CLIENTES 🔒
# ============================================================
@cargar_bp.route('/cargar/clientes', methods=['POST'])
@login_required
@solo_admin
def cargar_clientes():
    file = request.files.get('archivo')
    if not file or not allowed_file(file.filename):
        flash('Archivo inválido. Usa .xlsx', 'danger')
        return redirect(url_for('cargar.panel_cargar'))

    ruta = guardar_archivo(file)
    try:
        df = pd.read_excel(ruta)
        df.columns = [c.strip().lower() for c in df.columns]

        creados = actualizados = errores = 0
        for _, row in df.iterrows():
            try:
                nombre = str(row.get('nombre', '')).strip()
                cedula = str(row.get('cedula', '')).strip()
                if not nombre or not cedula or cedula == 'nan':
                    continue

                cliente = Cliente.query.filter_by(cedula=cedula).first()
                if cliente:
                    cliente.nombre    = nombre
                    cliente.telefono  = str(row.get('telefono', '') or '').strip()
                    cliente.direccion = str(row.get('direccion', '') or '').strip()
                    actualizados += 1
                else:
                    fecha_nac = None
                    fn_raw = row.get('fecha_nacimiento')
                    if pd.notna(fn_raw):
                        try:
                            fecha_nac = pd.to_datetime(fn_raw).date()
                        except:
                            pass

                    cliente = Cliente(
                        nombre           = nombre,
                        cedula           = cedula,
                        telefono         = str(row.get('telefono', '') or '').strip(),
                        direccion        = str(row.get('direccion', '') or '').strip(),
                        fecha_nacimiento = fecha_nac,
                        saldo_usd        = seguro_decimal(row.get('saldo_usd')),
                        saldo_bs         = seguro_decimal(row.get('saldo_bs')),
                        puntos           = int(row.get('puntos', 0) or 0),
                    )
                    db.session.add(cliente)
                    creados += 1
            except Exception:
                errores += 1

        db.session.commit()
        flash(f'✅ Clientes: {creados} creados, {actualizados} actualizados, {errores} errores.', 'success')
    except Exception as e:
        flash(f'❌ Error procesando archivo: {str(e)}', 'danger')

    return redirect(url_for('cargar.panel_cargar'))

# ============================================================
#   2. PROVEEDORES 🔒
# ============================================================
@cargar_bp.route('/cargar/proveedores', methods=['POST'])
@login_required
@solo_admin
def cargar_proveedores():
    file = request.files.get('archivo')
    if not file or not allowed_file(file.filename):
        flash('Archivo inválido. Usa .xlsx', 'danger')
        return redirect(url_for('cargar.panel_cargar'))

    ruta = guardar_archivo(file)
    try:
        df = pd.read_excel(ruta)
        df.columns = [c.strip().lower() for c in df.columns]

        creados = actualizados = errores = 0
        for _, row in df.iterrows():
            try:
                rif    = str(row.get('rif', '')).strip()
                nombre = str(row.get('nombre', '')).strip()
                if not rif or not nombre or rif == 'nan':
                    continue

                prov = Proveedor.query.filter_by(rif=rif).first()
                if prov:
                    prov.nombre            = nombre
                    prov.telefono          = str(row.get('telefono', '') or '').strip()
                    prov.direccion         = str(row.get('direccion', '') or '').strip()
                    prov.vendedor_nombre   = str(row.get('vendedor_nombre', '') or '').strip()
                    prov.vendedor_telefono = str(row.get('vendedor_telefono', '') or '').strip()
                    actualizados += 1
                else:
                    prov = Proveedor(
                        rif                  = rif,
                        nombre               = nombre,
                        telefono             = str(row.get('telefono', '') or '').strip(),
                        direccion            = str(row.get('direccion', '') or '').strip(),
                        vendedor_nombre      = str(row.get('vendedor_nombre', '') or '').strip(),
                        vendedor_telefono    = str(row.get('vendedor_telefono', '') or '').strip(),
                        saldo_pendiente_usd  = seguro_decimal(row.get('saldo_pendiente_usd')),
                    )
                    db.session.add(prov)
                    creados += 1
            except Exception:
                errores += 1

        db.session.commit()
        flash(f'✅ Proveedores: {creados} creados, {actualizados} actualizados, {errores} errores.', 'success')
    except Exception as e:
        flash(f'❌ Error procesando archivo: {str(e)}', 'danger')

    return redirect(url_for('cargar.panel_cargar'))

# ============================================================
#   3. INVENTARIO (PRODUCTOS) 🔒
# ============================================================
@cargar_bp.route('/cargar/inventario', methods=['POST'])
@login_required
@solo_admin
def cargar_inventario():
    file = request.files.get('archivo')
    if not file or not allowed_file(file.filename):
        flash('Archivo inválido. Usa .xlsx', 'danger')
        return redirect(url_for('cargar.panel_cargar'))

    ruta = guardar_archivo(file)
    try:
        df = pd.read_excel(ruta)
        df.columns = [c.strip().lower() for c in df.columns]

        creados = actualizados = errores = 0
        for _, row in df.iterrows():
            try:
                codigo = str(row.get('codigo', '')).strip()
                nombre = str(row.get('nombre', '')).strip()
                if not codigo or not nombre or codigo == 'nan':
                    continue

                prod = Producto.query.filter_by(codigo=codigo).first()
                if prod:
                    prod.nombre             = nombre
                    prod.categoria          = str(row.get('categoria', '') or '').strip()
                    prod.costo_usd          = seguro_decimal(row.get('costo_usd'))
                    prod.precio_normal_usd  = seguro_decimal(row.get('precio_normal_usd'))
                    prod.precio_oferta_usd  = seguro_decimal(row.get('precio_oferta_usd'))
                    antes_stock = prod.stock
                    prod.stock              = seguro_decimal(row.get('stock'))
                    prod.stock_minimo       = int(row.get('stock_minimo', 5) or 5)
                    actualizados += 1

                    # 📜 AUDITORIA
                    db.session.add(AuditoriaInventario(
                        usuario_id=current_user.id,
                        usuario_nombre=current_user.username,
                        producto_id=prod.id,
                        producto_nombre=prod.nombre,
                        accion='CARGA_EXCEL_STOCK_UPDATE',
                        cantidad_antes=antes_stock,
                        cantidad_despues=prod.stock,
                        fecha=datetime.now()
                    ))
                else:
                    prod = Producto(
                        codigo            = codigo,
                        nombre            = nombre,
                        categoria         = str(row.get('categoria', '') or '').strip(),
                        costo_usd         = seguro_decimal(row.get('costo_usd')),
                        precio_normal_usd = seguro_decimal(row.get('precio_normal_usd')),
                        precio_oferta_usd = seguro_decimal(row.get('precio_oferta_usd')),
                        stock             = seguro_decimal(row.get('stock')),
                        stock_minimo      = int(row.get('stock_minimo', 5) or 5),
                    )
                    db.session.add(prod)
                    db.session.flush()

                    # 📜 AUDITORIA
                    db.session.add(AuditoriaInventario(
                        usuario_id=current_user.id,
                        usuario_nombre=current_user.username,
                        producto_id=prod.id,
                        producto_nombre=prod.nombre,
                        accion='CARGA_EXCEL_NUEVO_PRODUCTO',
                        cantidad_antes=0,
                        cantidad_despues=prod.stock,
                        fecha=datetime.now()
                    ))
                    creados += 1
            except Exception:
                errores += 1

        db.session.commit()
        flash(f'✅ Inventario: {creados} creados, {actualizados} actualizados, {errores} errores.', 'success')
    except Exception as e:
        flash(f'❌ Error procesando archivo: {str(e)}', 'danger')

    return redirect(url_for('cargar.panel_cargar'))

# ============================================================
#   4. COMPRAS HISTÓRICAS 🔒
# ============================================================
@cargar_bp.route('/cargar/compras', methods=['POST'])
@login_required
@solo_admin
def cargar_compras():
    file = request.files.get('archivo')
    if not file or not allowed_file(file.filename):
        flash('Archivo inválido. Usa .xlsx', 'danger')
        return redirect(url_for('cargar.panel_cargar'))

    ruta = guardar_archivo(file)
    try:
        df = pd.read_excel(ruta)
        df.columns = [c.strip().lower() for c in df.columns]

        creados = errores = 0
        for _, row in df.iterrows():
            try:
                rif_prov = str(row.get('rif_proveedor', '')).strip()
                nro_fac  = str(row.get('numero_factura', '')).strip()
                if not rif_prov or rif_prov == 'nan':
                    continue

                prov = Proveedor.query.filter_by(rif=rif_prov).first()
                if not prov:
                    errores += 1
                    continue

                if nro_fac and Compra.query.filter_by(numero_factura=nro_fac).first():
                    continue

                fecha = datetime.utcnow()
                if pd.notna(row.get('fecha')):
                    try:
                        fecha = pd.to_datetime(row['fecha'])
                    except:
                        pass

                compra = Compra(
                    proveedor_id   = prov.id,
                    numero_factura = nro_fac,
                    total_usd      = seguro_decimal(row.get('total_usd')),
                    estado         = str(row.get('estado', 'Pendiente') or 'Pendiente').strip(),
                    metodo_pago    = str(row.get('metodo_pago', 'Credito') or 'Credito').strip(),
                    fecha          = fecha,
                )
                db.session.add(compra)
                db.session.flush()

                cod_prod = str(row.get('codigo_producto', '')).strip()
                if cod_prod and cod_prod != 'nan':
                    prod = Producto.query.filter_by(codigo=cod_prod).first()
                    if prod:
                        cantidad = Decimal(str(row.get('cantidad', 1) or 1))
                        costo    = Decimal(str(row.get('costo_unitario', 0) or 0))
                        detalle  = CompraDetalle(
                            compra_id      = compra.id,
                            producto_id    = prod.id,
                            cantidad       = cantidad,
                            costo_unitario = costo,
                        )
                        db.session.add(detalle)
                        antes_compra = prod.stock
                        prod.stock += cantidad

                        # 📜 AUDITORIA
                        db.session.add(AuditoriaInventario(
                            usuario_id=current_user.id,
                            usuario_nombre=current_user.username,
                            producto_id=prod.id,
                            producto_nombre=prod.nombre,
                            accion='CARGA_EXCEL_COMPRA_HISTORICA',
                            cantidad_antes=antes_compra,
                            cantidad_despues=prod.stock,
                            fecha=datetime.now()
                        ))

                creados += 1
            except Exception:
                errores += 1

        db.session.commit()
        flash(f'✅ Compras: {creados} creadas, {errores} errores.', 'success')
    except Exception as e:
        flash(f'❌ Error procesando archivo: {str(e)}', 'danger')

    return redirect(url_for('cargar.panel_cargar'))

# ============================================================
#   5. VENTAS HISTÓRICAS 🔒
# ============================================================
@cargar_bp.route('/cargar/ventas', methods=['POST'])
@login_required
@solo_admin
def cargar_ventas():
    file = request.files.get('archivo')
    if not file or not allowed_file(file.filename):
        flash('Archivo inválido. Usa .xlsx', 'danger')
        return redirect(url_for('cargar.panel_cargar'))

    ruta = guardar_archivo(file)
    try:
        df = pd.read_excel(ruta)
        df.columns = [c.strip().lower() for c in df.columns]

        tasa_obj = TasaBCV.query.order_by(TasaBCV.id.desc()).first()
        tasa_default = float(tasa_obj.valor) if tasa_obj else 1.0

        creados = errores = 0
        for _, row in df.iterrows():
            try:
                total_usd = Decimal(str(row.get('total_usd', 0) or 0))
                if total_usd <= 0:
                    continue

                cliente_id = None
                cedula = str(row.get('cedula_cliente', '')).strip()
                if cedula and cedula != 'nan':
                    cli = Cliente.query.filter_by(cedula=cedula).first()
                    if cli:
                        cliente_id = cli.id

                fecha = datetime.utcnow()
                if pd.notna(row.get('fecha')):
                    try:
                        fecha = pd.to_datetime(row['fecha'])
                    except:
                        pass

                es_fiado = str(row.get('es_fiado', 'no')).strip().lower() in ('si', 'sí', '1', 'true', 'yes')

                venta = Venta(
                    fecha                 = fecha,
                    cliente_id            = cliente_id,
                    total_usd             = total_usd,
                    tasa_momento          = Decimal(str(row.get('tasa_momento', tasa_default) or tasa_default)),
                    es_fiado              = es_fiado,
                    pagada                = not es_fiado,
                    pago_efectivo_usd     = Decimal(str(row.get('pago_efectivo_usd', 0) or 0)),
                    pago_efectivo_bs      = Decimal(str(row.get('pago_efectivo_bs', 0) or 0)),
                    pago_movil_bs         = Decimal(str(row.get('pago_movil_bs', 0) or 0)),
                    pago_transferencia_bs = Decimal(str(row.get('pago_transferencia_bs', 0) or 0)),
                    biopago_bdv           = Decimal(str(row.get('biopago_bdv', 0) or 0)),
                )
                db.session.add(venta)
                db.session.flush()

                cod_prod = str(row.get('codigo_producto', '')).strip()
                if cod_prod and cod_prod != 'nan':
                    prod = Producto.query.filter_by(codigo=cod_prod).first()
                    if prod:
                        cantidad = Decimal(str(row.get('cantidad', 1) or 1))
                        precio   = Decimal(str(row.get('precio_unitario_usd', 0) or 0))
                        detalle  = DetalleVenta(
                            venta_id            = venta.id,
                            producto_id         = prod.id,
                            cantidad            = cantidad,
                            precio_unitario_usd = precio,
                        )
                        db.session.add(detalle)
                        antes_venta = prod.stock
                        prod.stock = max(0, prod.stock - cantidad)

                        # 📜 AUDITORIA
                        db.session.add(AuditoriaInventario(
                            usuario_id=current_user.id,
                            usuario_nombre=current_user.username,
                            producto_id=prod.id,
                            producto_nombre=prod.nombre,
                            accion='CARGA_EXCEL_VENTA_HISTORICA',
                            cantidad_antes=antes_venta,
                            cantidad_despues=prod.stock,
                            fecha=datetime.now()
                        ))

                if es_fiado and cliente_id:
                    cli = Cliente.query.get(cliente_id)
                    if cli:
                        cli.saldo_usd = (cli.saldo_usd or Decimal('0')) + total_usd

                creados += 1
            except Exception:
                errores += 1

        db.session.commit()
        flash(f'✅ Ventas: {creados} creadas, {errores} errores.', 'success')
    except Exception as e:
        flash(f'❌ Error procesando archivo: {str(e)}', 'danger')

    return redirect(url_for('cargar.panel_cargar'))

# ============================================================
#   6. SALDOS DE FIADO 🔒
# ============================================================
@cargar_bp.route('/cargar/fiado', methods=['POST'])
@login_required
@solo_admin
def cargar_fiado():
    file = request.files.get('archivo')
    if not file or not allowed_file(file.filename):
        flash('Archivo inválido. Usa .xlsx', 'danger')
        return redirect(url_for('cargar.panel_cargar'))

    ruta = guardar_archivo(file)
    try:
        df = pd.read_excel(ruta)
        df.columns = [c.strip().lower() for c in df.columns]

        actualizados = errores = no_encontrados = 0
        for _, row in df.iterrows():
            try:
                cedula = str(row.get('cedula', '')).strip()
                if not cedula or cedula == 'nan':
                    continue

                cli = Cliente.query.filter_by(cedula=cedula).first()
                if not cli:
                    no_encontrados += 1
                    continue

                cli.saldo_usd = Decimal(str(row.get('saldo_usd', 0) or 0))
                cli.saldo_bs  = Decimal(str(row.get('saldo_bs',  0) or 0))
                actualizados += 1
            except Exception:
                errores += 1

        db.session.commit()
        flash(f'✅ Fiado: {actualizados} actualizados, {no_encontrados} no encontrados, {errores} errores.', 'success')
    except Exception as e:
        flash(f'❌ Error procesando archivo: {str(e)}', 'danger')

    return redirect(url_for('cargar.panel_cargar'))

# ============================================================
#   7. CUENTAS POR PAGAR 🔒
# ============================================================
@cargar_bp.route('/cargar/cuentas_pagar', methods=['POST'])
@login_required
@solo_admin
def cargar_cuentas_pagar():
    file = request.files.get('archivo')
    if not file or not allowed_file(file.filename):
        flash('Archivo inválido. Usa .xlsx', 'danger')
        return redirect(url_for('cargar.panel_cargar'))

    ruta = guardar_archivo(file)
    try:
        df = pd.read_excel(ruta)
        df.columns = [c.strip().lower() for c in df.columns]

        creados = errores = no_encontrados = 0
        for _, row in df.iterrows():
            try:
                rif_prov = str(row.get('rif_proveedor', '')).strip()
                nro_fac  = str(row.get('numero_factura', '')).strip()
                if not rif_prov or rif_prov == 'nan':
                    continue

                prov = Proveedor.query.filter_by(rif=rif_prov).first()
                if not prov:
                    no_encontrados += 1
                    continue

                compra = None
                if nro_fac and nro_fac != 'nan':
                    compra = Compra.query.filter_by(numero_factura=nro_fac).first()

                if not compra:
                    compra = Compra(
                        proveedor_id   = prov.id,
                        numero_factura = nro_fac,
                        total_usd      = Decimal(str(row.get('monto_total_usd', 0) or 0)),
                        estado         = 'Pendiente',
                        metodo_pago    = 'Credito',
                    )
                    db.session.add(compra)
                    db.session.flush()

                monto_total   = Decimal(str(row.get('monto_total_usd', 0) or 0))
                monto_abonado = Decimal(str(row.get('monto_abonado_usd', 0) or 0))
                saldo         = monto_total - monto_abonado

                estatus = 'Pendiente'
                if saldo <= 0:
                    estatus = 'Pagado'
                elif monto_abonado > 0:
                    estatus = 'Parcial'

                fecha_emision = datetime.utcnow()
                if pd.notna(row.get('fecha_emision')):
                    try:
                        fecha_emision = pd.to_datetime(row['fecha_emision'])
                    except:
                        pass

                cpp = CuentaPorPagar(
                    proveedor_id        = prov.id,
                    compra_id           = compra.id,
                    numero_factura      = nro_fac,
                    fecha_emision       = fecha_emision,
                    monto_total_usd     = monto_total,
                    monto_abonado_usd   = monto_abonado,
                    saldo_pendiente_usd = saldo,
                    estatus             = estatus,
                )
                db.session.add(cpp)
                prov.saldo_pendiente_usd = (prov.saldo_pendiente_usd or Decimal('0')) + saldo
                creados += 1
            except Exception:
                errores += 1

        db.session.commit()
        flash(f'✅ Cuentas por Pagar: {creados} creadas, {no_encontrados} no encontrados, {errores} errores.', 'success')
    except Exception as e:
        flash(f'❌ Error procesando archivo: {str(e)}', 'danger')

    return redirect(url_for('cargar.panel_cargar'))

# ============================================================
#   8. SALDOS DE CAJA (CAPITAL INICIAL) 🔒  👈 NUEVO
# ============================================================
@cargar_bp.route('/cargar/saldos_caja', methods=['POST'])
@login_required
@solo_admin
def cargar_saldos_caja():
    file = request.files.get('archivo')
    if not file or not allowed_file(file.filename):
        flash('Archivo inválido. Usa .xlsx', 'danger')
        return redirect(url_for('cargar.panel_cargar'))

    ruta = guardar_archivo(file)
    try:
        df = pd.read_excel(ruta)
        df.columns = [c.strip().lower() for c in df.columns]

        creados = errores = 0
        for _, row in df.iterrows():
            try:
                tipo_caja = str(row.get('tipo_caja', '')).strip()
                if not tipo_caja or tipo_caja == 'nan':
                    continue

                monto_usd   = seguro_decimal(row.get('monto_usd'))
                monto_bs    = seguro_decimal(row.get('monto_bs'))
                descripcion = str(row.get('descripcion', '') or '').strip()
                if not descripcion:
                    descripcion = f"Saldo inicial cargado por Excel en {tipo_caja}"

                mov = MovimientoCaja(
                    tipo_movimiento = 'INGRESO',
                    categoria       = 'SALDO INICIAL',
                    tipo_caja       = tipo_caja,
                    monto_usd       = monto_usd,
                    monto_bs        = monto_bs,
                    descripcion     = descripcion,
                    usuario_id      = current_user.id,
                    fecha           = datetime.now(),
                )
                db.session.add(mov)

                # ============================================================
                # 📒 ASIENTO CONTABLE DOBLE ENTRADA
                # DEBE: Cuenta de Caja correspondiente
                # HABER: Capital Social 3.1.01
                # ============================================================
                mapa_cuentas = {
                    'Efectivo USD':   '1.1.01.01',
                    'Efectivo Bs':    '1.1.01.02',
                    'Pago Móvil':     '1.1.01.03',
                    'Transferencia':  '1.1.01.03',
                    'Biopago':        '1.1.01.04',
                    'Punto de Venta': '1.1.01.05',
                }
                cuenta_caja = mapa_cuentas.get(tipo_caja, '1.1.01.01')
                es_usd = tipo_caja == 'Efectivo USD'

                registrar_asiento(
                    descripcion=descripcion,
                    tasa=1.0,
                    referencia_tipo='CAPITAL_INICIAL',
                    referencia_id=0,
                    movimientos=[
                        {
                            'cuenta_codigo': cuenta_caja,
                            'debe_usd':  float(monto_usd) if es_usd else 0,
                            'haber_usd': 0,
                            'debe_bs':   float(monto_bs) if not es_usd else 0,
                            'haber_bs':  0,
                        },
                        {
                            'cuenta_codigo': '3.1.01',
                            'debe_usd':  0,
                            'haber_usd': float(monto_usd) if es_usd else 0,
                            'debe_bs':   0,
                            'haber_bs':  float(monto_bs) if not es_usd else 0,
                        },
                    ]
                )

                creados += 1
            except Exception as e:
                errores += 1

        db.session.commit()
        flash(f'✅ Saldos de Caja: {creados} registros cargados, {errores} errores.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error procesando archivo: {str(e)}', 'danger')

    return redirect(url_for('cargar.panel_cargar'))

# ============================================================
#   DESCARGAR PLANTILLAS
# ============================================================
@cargar_bp.route('/cargar/plantilla/<modulo>')
def descargar_plantilla(modulo):
    plantillas = {
        'clientes':      ['nombre', 'cedula', 'telefono', 'direccion', 'fecha_nacimiento', 'saldo_usd', 'saldo_bs', 'puntos'],
        'proveedores':   ['rif', 'nombre', 'telefono', 'direccion', 'vendedor_nombre', 'vendedor_telefono', 'saldo_pendiente_usd'],
        'inventario':    ['codigo', 'nombre', 'categoria', 'costo_usd', 'precio_normal_usd', 'precio_oferta_usd', 'stock', 'stock_minimo'],
        'compras':       ['rif_proveedor', 'numero_factura', 'fecha', 'total_usd', 'estado', 'metodo_pago', 'codigo_producto', 'cantidad', 'costo_unitario'],
        'ventas':        ['fecha', 'cedula_cliente', 'total_usd', 'tasa_momento', 'es_fiado', 'pago_efectivo_usd', 'pago_efectivo_bs', 'pago_movil_bs', 'codigo_producto', 'cantidad', 'precio_unitario_usd'],
        'fiado':         ['cedula', 'nombre', 'saldo_usd', 'saldo_bs'],
        'cuentas_pagar': ['rif_proveedor', 'numero_factura', 'fecha_emision', 'monto_total_usd', 'monto_abonado_usd'],
        'saldos_caja':   ['tipo_caja', 'monto_usd', 'monto_bs', 'descripcion'],  # 👈 NUEVO
    }

    if modulo not in plantillas:
        flash('Módulo no encontrado', 'danger')
        return redirect(url_for('cargar.panel_cargar'))

    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = modulo.capitalize()

    columnas = plantillas[modulo]
    for col_idx, col_name in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(len(col_name) + 4, 15)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'plantilla_{modulo}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )