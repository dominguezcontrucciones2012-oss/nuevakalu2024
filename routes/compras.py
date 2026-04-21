from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from routes.decorators import staff_required
from models import db, Producto, Proveedor, Compra, CompraDetalle, CuentaPorPagar, AbonoCuentaPorPagar, MovimientoCaja, TasaBCV, AuditoriaInventario
import logging
from decimal import Decimal
from datetime import datetime
from utils import seguro_decimal
import io
import openpyxl

compras_bp = Blueprint('compras', __name__)
logger = logging.getLogger('KALU.compras')

# ==========================================================
#   LISTA DE COMPRAS + CARGA RÁPIDA
# ==========================================================
@compras_bp.route('/compras')
@login_required
@staff_required
def lista_compras():
    compras     = Compra.query.order_by(Compra.fecha.desc()).all()
    proveedores = Proveedor.query.order_by(Proveedor.nombre).all()
    return render_template('compras.html', compras=compras, proveedores=proveedores)


# ==========================================================
#   BUSCAR PRODUCTO POR CÓDIGO (Para el escáner)
# ==========================================================
@compras_bp.route('/buscar_producto/<codigo>')
@login_required
@staff_required
def buscar_producto(codigo):
    prod = Producto.query.filter_by(codigo=codigo).first()
    if prod:
        return jsonify({
            'id':     prod.id,
            'codigo': prod.codigo,
            'nombre': prod.nombre,
            'costo':  str(seguro_decimal(prod.costo_usd))
        })
    return jsonify({'id': None})


# ==========================================================
#   BUSCAR PRODUCTO POR NOMBRE (búsqueda parcial)
# ==========================================================
@compras_bp.route('/buscar_producto_nombre/<texto>')
@login_required
@staff_required
def buscar_producto_nombre(texto):
    texto = texto.strip()
    if not texto:
        return jsonify([])
    productos = Producto.query.filter(Producto.nombre.ilike(f'%{texto}%')).limit(10).all()
    resultados = []
    for p in productos:
        resultados.append({
            'id':     p.id,
            'codigo': p.codigo,
            'nombre': p.nombre,
            'costo':  str(seguro_decimal(p.costo_usd))
        })
    return jsonify(resultados)


# ==========================================================
#   CREAR PRODUCTO RÁPIDO (Producto nuevo desde compras)
# ==========================================================
@compras_bp.route('/crear_producto_rapido', methods=['POST'])
@login_required
@staff_required
def crear_producto_rapido():
    try:
        data          = request.get_json()
        codigo        = data.get('codigo', '').strip()
        nombre        = data.get('nombre', '').strip()
        costo         = seguro_decimal(data.get('costo'))
        precio        = seguro_decimal(data.get('precio'))
        precio_oferta = seguro_decimal(data.get('precio_oferta'))
        stock_inicial = seguro_decimal(data.get('stock'))

        if not nombre:
            return jsonify({'id': None, 'message': 'Falta el nombre'}), 400

        existente = Producto.query.filter_by(codigo=codigo).first()
        if existente:
            return jsonify({
                'id':     existente.id,
                'codigo': existente.codigo,
                'nombre': existente.nombre,
                'costo':  str(existente.costo_usd or Decimal('0.00'))
            })

        nuevo = Producto(
            codigo            = codigo,
            nombre            = nombre,
            costo_usd         = costo,
            precio_normal_usd = precio,
            precio_oferta_usd = precio_oferta if precio_oferta > 0 else precio,
            stock             = stock_inicial
        )
        db.session.add(nuevo)
        db.session.commit()

        return jsonify({
            'id':     nuevo.id,
            'codigo': nuevo.codigo,
            'nombre': nuevo.nombre,
            'costo':  str(nuevo.costo_usd.quantize(Decimal('0.01'))) if nuevo.costo_usd else "0.00"
        })

    except Exception as e:
        db.session.rollback()
        logger.error(f"❌ Error crear_producto_rapido: {e}")
        return jsonify({'id': None, 'message': str(e)}), 500


# ==========================================================
#   PROCESAR COMPRA RÁPIDA (Guardar factura + subir stock)
# ==========================================================
@compras_bp.route('/procesar_compra_rapida', methods=['POST'])
@login_required
@staff_required
def procesar_compra_rapida():
    try:
        data         = request.get_json()
        proveedor_id = data.get('proveedor_id')
        num_factura  = data.get('numero_factura', '').strip()
        items        = data.get('items', [])
        metodo_pago  = data.get('metodo_pago', 'Credito')
        caja_origen  = data.get('caja_origen', None)

        if not proveedor_id or not items:
            return jsonify({'status': 'error', 'message': 'Faltan datos'}), 400

        # 1. Calcular total
        # ✅ CORREGIDO: Decimal en vez de int para que 1.5 kilos no se convierta en 1
        total_usd = Decimal('0.00')
        for i in items:
            total_usd += Decimal(str(i['costo'])) * Decimal(str(i['cantidad']))

        # 2. Crear la Compra
        nueva_compra = Compra(
            proveedor_id   = proveedor_id,
            numero_factura = num_factura,
            fecha          = datetime.now(),
            total_usd      = total_usd,
            metodo_pago    = metodo_pago,
            caja_origen    = caja_origen,
            estado         = 'Pagado' if metodo_pago != 'Credito' else 'Pendiente'
        )
        db.session.add(nueva_compra)
        db.session.flush()

        # 3. Actualizar stock y costos de productos
        for item in items:
            prod = Producto.query.get(item['id'])
            if prod:
                # ✅ CORREGIDO: Decimal para que 20.5 kilos no se convierta en 20
                cantidad_item  = Decimal(str(item['cantidad']))
                prod.stock     = (prod.stock or Decimal('0')) + cantidad_item
                prod.costo_usd = Decimal(str(item['costo']))
                detalle = CompraDetalle(
                    compra_id      = nueva_compra.id,
                    producto_id    = prod.id,
                    cantidad       = cantidad_item,
                    costo_unitario = Decimal(str(item['costo']))
                )
                db.session.add(detalle)

                # 📜 AUDITORIA
                antes = prod.stock - cantidad_item
                db.session.add(AuditoriaInventario(
                    usuario_id      = current_user.id,
                    usuario_nombre  = current_user.username,
                    producto_id     = prod.id,
                    producto_nombre = prod.nombre,
                    accion          = 'COMPRA_MERCANCIA',
                    cantidad_antes  = antes,
                    cantidad_despues = prod.stock,
                    fecha           = datetime.now()
                ))

        # 4. Lógica financiera según método de pago
        if metodo_pago in ('Contado USD', 'Contado Bs'):
            mov = MovimientoCaja(
                tipo_caja       = caja_origen,
                tipo_movimiento = 'EGRESO',
                categoria       = 'Compra de Mercancía',
                monto           = total_usd,
                descripcion     = f'Pago contado factura {num_factura}',
                modulo_origen   = 'Compra',
                referencia_id   = nueva_compra.id
            )
            db.session.add(mov)
        else:
            nueva_cxp = CuentaPorPagar(
                proveedor_id        = proveedor_id,
                compra_id           = nueva_compra.id,
                numero_factura      = num_factura,
                monto_total_usd     = total_usd,
                saldo_pendiente_usd = total_usd
            )
            db.session.add(nueva_cxp)

            prov = Proveedor.query.get(proveedor_id)
            prov.saldo_pendiente_usd = (prov.saldo_pendiente_usd or Decimal('0.00')) + total_usd

        # 🔄 SINCRONIZACIÓN CON LIBRETA PRODUCTOR (Si el proveedor es productor)
        prov = Proveedor.query.get(proveedor_id)
        if prov and prov.es_productor:
            from models import MovimientoProductor
            debe = total_usd if metodo_pago in ('Contado USD', 'Contado Bs') else Decimal('0.00')
            db.session.add(MovimientoProductor(
                proveedor_id=prov.id,
                tipo='COMPRA_MERCANCIA',
                descripcion=f"Compra de mercancía (Fac: {num_factura}) - Modulo Compras",
                haber=total_usd,
                debe=debe,
                saldo_momento=prov.saldo_pendiente_usd,
            ))

        db.session.commit()
        return jsonify({'status': 'success'})

    except Exception as e:
        db.session.rollback()
        logger.error(f"❌ Error procesar_compra_rapida: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ==========================================================
#   ABONAR A PROVEEDOR (Bajar deuda de una factura)
# ==========================================================
@compras_bp.route('/compras/abonar', methods=['POST'])
@login_required
@staff_required
def abonar_proveedor():
    try:
        data        = request.get_json()
        cxp_id      = data.get('cxp_id')
        caja_origen = data.get('caja_origen')
        moneda      = data.get('moneda', 'USD')
        tasa        = Decimal(str(data.get('tasa_bcv', 1) or 1))

        # ✅ FIX: El frontend envía monto_usd (ya convertido) y monto_real (monto crudo ingresado)
        monto_usd   = Decimal(str(data.get('monto_usd', 0)))
        monto_real  = Decimal(str(data.get('monto_real', monto_usd)))

        # 🛡️ ROBUSTEZ: Si moneda es Bs y el frontend mandó USD == BS (error de tasa 1), recalculamos
        if moneda == 'Bs' and abs(monto_usd - monto_real) < 0.01 and tasa > 1:
            monto_usd = (monto_real / tasa).quantize(Decimal('0.01'))
            logger.info(f"🔄 Recalculado monto_usd de seguridad: {monto_real} Bs -> {monto_usd} USD (Tasa: {tasa})")

        cxp = CuentaPorPagar.query.get(cxp_id)
        if not cxp or monto_usd <= 0:
            return jsonify({'status': 'error', 'message': 'Datos inválidos o monto en cero'})

        # ✅ VALIDACIÓN DE SEGURIDAD REPARADA
        # 🛑 BLOQUEO: Verificar que la caja tiene saldo suficiente
        ingresos = db.session.query(db.func.sum(MovimientoCaja.monto))\
            .filter_by(tipo_caja=caja_origen, tipo_movimiento='INGRESO').scalar() or Decimal('0')
        egresos = db.session.query(db.func.sum(MovimientoCaja.monto))\
            .filter_by(tipo_caja=caja_origen, tipo_movimiento='EGRESO').scalar() or Decimal('0')
        saldo_caja = Decimal(str(ingresos)) - Decimal(str(egresos))

        if monto_real > saldo_caja:
            return jsonify({
                'status': 'error',
                'message': f'❌ Saldo insuficiente en {caja_origen}. Disponible: {saldo_caja.quantize(Decimal("0.01"))}, necesitas: {monto_real.quantize(Decimal("0.01"))}'
            })

        if monto_usd > cxp.saldo_pendiente_usd + Decimal('0.05'):
            return jsonify({
                'status': 'error',
                'message': f'El monto (${monto_usd}) supera el saldo pendiente (${cxp.saldo_pendiente_usd})'
            })

        nuevo_abono = AbonoCuentaPorPagar(
            cuenta_id   = cxp.id,
            monto_usd   = monto_usd,
            metodo_pago = caja_origen,
            descripcion = f'Abono Fac {cxp.numero_factura} | '
                          f'{"Bs " + str(monto_real) + " @ " + str(tasa) if moneda == "Bs" else "$" + str(monto_usd)}'
        )

        db.session.add(nuevo_abono)

        cxp.monto_abonado_usd   += monto_usd
        cxp.saldo_pendiente_usd -= monto_usd
        
        if cxp.saldo_pendiente_usd <= 0:
            cxp.saldo_pendiente_usd = Decimal('0.00')
            cxp.estatus = 'Pagado'
        else:
            cxp.estatus = 'Parcial'

        prov = Proveedor.query.get(cxp.proveedor_id)
        prov.saldo_pendiente_usd = (prov.saldo_pendiente_usd or Decimal('0.00')) - monto_usd
        if prov.saldo_pendiente_usd < 0:
            prov.saldo_pendiente_usd = Decimal('0.00')

        # 💰 REGISTRO EN CAJA (Siempre como EGRESO para abonos)
        # La descripción ahora es bilingüe para mayor claridad del dueño
        desc_mov = f'Abono Fac {cxp.numero_factura} ({moneda}) | Ref: ${monto_usd:.2f} / {monto_real:,.2f} Bs'
        
        mov = MovimientoCaja(
            tipo_caja       = caja_origen,
            tipo_movimiento = 'EGRESO',
            categoria       = 'Pago a Proveedor',
            monto           = monto_real,
            tasa_dia        = tasa,
            descripcion     = desc_mov,
            modulo_origen   = 'CXP/Compras',
            referencia_id   = cxp.id,
            user_id         = current_user.id
        )

        db.session.add(mov)

        # 🔄 SINCRONIZACIÓN CON LIBRETA PRODUCTOR (Si el proveedor es productor)
        if prov and prov.es_productor:
            from models import MovimientoProductor
            db.session.add(MovimientoProductor(
                proveedor_id=prov.id,
                tipo='PAGO_CXP',
                descripcion=f"Abono a factura {cxp.numero_factura} desde CXP",
                debe=monto_usd,
                haber=0,
                saldo_momento=prov.saldo_pendiente_usd
            ))

        db.session.commit()
        return jsonify({
            'status':     'success',
            'monto_usd':  str(monto_usd.quantize(Decimal('0.01'))),
            'monto_real': str(monto_real.quantize(Decimal('0.01'))),
            'moneda':     moneda
        })

    except Exception as e:
        db.session.rollback()
        logger.error(f"❌ Error abonar_proveedor: {e}")
        return jsonify({'status': 'error', 'message': str(e)})
# ==========================================================
#   DESCARGAR PLANTILLA EXCEL
# ==========================================================
@compras_bp.route('/inventario/plantilla_excel')
@login_required
@staff_required
def descargar_plantilla():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    # ✅ CORREGIDO: Agregada columna unidad_medida en la plantilla
    ws.append(["codigo", "nombre", "costo_usd", "precio_normal_usd", "precio_oferta_usd", "stock", "unidad_medida"])
    ws.append(["00001", "Ejemplo Harina", 1.50, 2.00, 1.80, 45.5, "KG"])
    ws.append(["00002", "Ejemplo Refresco", 0.50, 0.80, 0.70, 24, "UND"])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="plantilla_inventario.xlsx",
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ==========================================================
#   CARGAR INVENTARIO DESDE EXCEL
# ==========================================================
@compras_bp.route('/inventario/cargar_excel', methods=['POST'])
@login_required
@staff_required
def cargar_excel():
    archivo = request.files.get('archivo_excel')
    if not archivo:
        return jsonify({'status': 'error', 'message': 'No se recibió archivo'})
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        creados = 0
        actualizados = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            # ✅ CORREGIDO: Lee unidad_medida del Excel (columna 7, opcional)
            codigo, nombre, costo, precio, oferta, stock, *extra = row
            unidad = extra[0] if extra else 'UND'
            if not nombre:
                continue
            prod = Producto.query.filter_by(codigo=str(codigo)).first()
            if prod:
                prod.costo_usd         = Decimal(str(costo)) if costo is not None else Decimal('0')
                prod.precio_normal_usd = Decimal(str(precio)) if precio is not None else Decimal('0')
                prod.precio_oferta_usd = Decimal(str(oferta or precio or 0))
                # ✅ CORREGIDO: Decimal en vez de int para kilos
                antes_stock = prod.stock
                prod.stock             = Decimal(str(stock or 0))
                prod.unidad_medida     = str(unidad or 'UND').strip().upper()
                actualizados += 1

                # 📜 AUDITORIA
                db.session.add(AuditoriaInventario(
                    usuario_id      = current_user.id,
                    usuario_nombre  = current_user.username,
                    producto_id     = prod.id,
                    producto_nombre = prod.nombre,
                    accion          = 'CARGA_EXCEL_UPDATE',
                    cantidad_antes  = antes_stock,
                    cantidad_despues = prod.stock,
                    fecha           = datetime.now()
                ))
            else:
                nuevo = Producto(
                    codigo            = str(codigo),
                    nombre            = str(nombre),
                    costo_usd         = Decimal(str(costo or 0)),
                    precio_normal_usd = Decimal(str(precio or 0)),
                    precio_oferta_usd = Decimal(str(oferta or precio or 0)),
                    stock             = Decimal(str(stock or 0)),
                    unidad_medida     = str(unidad or 'UND').strip().upper()
                )
                db.session.add(nuevo)
                db.session.flush() # Para obtener el ID

                # 📜 AUDITORIA
                db.session.add(AuditoriaInventario(
                    usuario_id      = current_user.id,
                    usuario_nombre  = current_user.username,
                    producto_id     = nuevo.id,
                    producto_nombre = nuevo.nombre,
                    accion          = 'CARGA_EXCEL_NUEVO',
                    cantidad_antes  = 0,
                    cantidad_despues = nuevo.stock,
                    fecha           = datetime.now()
                ))
                creados += 1
        db.session.commit()
        return jsonify({'status': 'success', 'creados': creados, 'actualizados': actualizados})
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)})


# ==========================================================
#   VISTA CUENTAS POR PAGAR
# ==========================================================
@compras_bp.route('/cuentas_por_pagar')
@login_required
@staff_required
def cuentas_por_pagar():
    cuentas     = CuentaPorPagar.query.order_by(CuentaPorPagar.fecha.desc()).all()
    proveedores = Proveedor.query.order_by(Proveedor.nombre).all()
    tasa_obj    = TasaBCV.query.order_by(TasaBCV.fecha.desc()).first()
    tasa_bcv    = Decimal(str(tasa_obj.valor)) if tasa_obj else Decimal('1.00')
    return render_template('cuentas_por_pagar.html',
                           cuentas=cuentas,
                           proveedores=proveedores,
                           tasa_bcv=tasa_bcv)


@compras_bp.route('/compras/<int:compra_id>/detalle')
@login_required
@staff_required
def detalle_compra(compra_id):
    compra = Compra.query.get_or_404(compra_id)
    return render_template('detalle_compra.html', compra=compra)